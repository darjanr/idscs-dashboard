#!/usr/bin/env python3
"""Regenerate the locale + map files from the edited translations.xlsx.

Reads translations.xlsx (the single source of truth the client edits) and writes:
  - src/i18n/mk.json, al.json, en.json   (UI strings)
  - src/i18n/parties.json                (party names + acronyms)
  - src/i18n/mp_names.json               (MP name translations)

After running, rebuild the site (npm run build). Round-trips with i18n_export.py.

Usage:  python3 scripts/i18n_import.py
"""
import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
I18N = ROOT / "src" / "i18n"
XLSX = ROOT / "translations.xlsx"


def set_path(root, parts, value):
    """Set a dot-path into nested dict/list, creating containers as needed."""
    cur = root
    for i, p in enumerate(parts):
        last = i == len(parts) - 1
        key = int(p) if p.isdigit() else p
        if last:
            if isinstance(key, int):
                while len(cur) <= key:
                    cur.append(None)
                cur[key] = value
            else:
                cur[key] = value
            return
        child_is_list = parts[i + 1].isdigit()
        if isinstance(key, int):
            while len(cur) <= key:
                cur.append(None)
            if cur[key] is None:
                cur[key] = [] if child_is_list else {}
            cur = cur[key]
        else:
            if cur.get(key) is None:
                cur[key] = [] if child_is_list else {}
            cur = cur[key]


def cell(v):
    return "" if v is None else str(v)


def dump(obj, path):
    path.write_text(json.dumps(obj, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def main():
    if not XLSX.exists():
        raise SystemExit(f"Missing {XLSX}. Run scripts/i18n_export.py first, then edit it.")
    wb = openpyxl.load_workbook(XLSX)

    # UI text -> mk/al/en.json
    ws = wb["UI text"]
    mk, al, en = {}, {}, {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = row[0]
        if not key:
            continue
        parts = str(key).split(".")
        set_path(mk, parts, cell(row[1]))
        set_path(al, parts, cell(row[2]))
        set_path(en, parts, cell(row[3]))
    dump(mk, I18N / "mk.json")
    dump(al, I18N / "al.json")
    dump(en, I18N / "en.json")

    # Parties
    wp = wb["Parties"]
    parties = {}
    for row in wp.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        parties[str(row[0])] = {
            "al": cell(row[1]), "en": cell(row[2]),
            "acrMk": cell(row[3]), "acrAl": cell(row[4]), "acrEn": cell(row[5]),
        }
    dump(parties, I18N / "parties.json")

    # MP names
    wn = wb["MP names"]
    names = {}
    for row in wn.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        names[str(row[0])] = {"al": cell(row[1]), "en": cell(row[2])}
    dump(names, I18N / "mp_names.json")

    # Institutions + Case types (simple MK | AL | EN maps)
    counts = {}
    for sheet_name, fname in [("Institutions", "institutions.json"), ("Case types", "case_types.json")]:
        if sheet_name not in wb.sheetnames:
            continue
        data = {}
        for row in wb[sheet_name].iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            data[str(row[0])] = {"al": cell(row[1]), "en": cell(row[2])}
        dump(data, I18N / fname)
        counts[sheet_name] = len(data)

    # Questions + answers -> public/data/questions_i18n.json (only non-empty rows)
    if "Questions" in wb.sheetnames:
        qi18n = {}
        for row in wb["Questions"].iter_rows(min_row=2, values_only=True):
            qid = row[0]
            if not qid:
                continue
            entry = {}
            if cell(row[2]) or cell(row[3]):
                entry["q"] = {"al": cell(row[2]), "en": cell(row[3])}
            if cell(row[5]) or cell(row[6]):
                entry["a"] = {"al": cell(row[5]), "en": cell(row[6])}
            if entry:
                qi18n[str(qid)] = entry
        dump(qi18n, ROOT / "public" / "data" / "questions_i18n.json")
        counts["Questions"] = len(qi18n)

    print(f"✓ regenerated locales (UI: {len(_flat(mk))} keys), parties: {len(parties)}, "
          f"MP names: {len(names)}, "
          + ", ".join(f"{k}: {v}" for k, v in counts.items()))


def _flat(obj, n=0):
    if isinstance(obj, dict):
        return [x for v in obj.values() for x in _flat(v)]
    if isinstance(obj, list):
        return [x for v in obj for x in _flat(v)]
    return [obj]


if __name__ == "__main__":
    main()
