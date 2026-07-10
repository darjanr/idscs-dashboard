#!/usr/bin/env python3
"""Export all translatable text into a single review workbook (translations.xlsx).

Sheets:
  - "UI text"  : every interface string, columns key | MK | AL | EN
  - "Parties"  : party names + acronyms, MK | AL | EN | Acr MK/AL/EN
  - "MP names" : every MP name, MK | AL | EN

The workbook is the editable single source of truth. After the client edits it,
run scripts/i18n_import.py to regenerate the locale + map files the site uses.

Usage:  python3 scripts/i18n_export.py
"""
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parent.parent
I18N = ROOT / "src" / "i18n"
OUT = ROOT / "translations.xlsx"


def load(p):
    return json.loads(p.read_text(encoding="utf-8"))


def flatten(obj, prefix=""):
    """Flatten nested dicts/lists to dot-path leaves (lists use numeric segments)."""
    items = []
    if isinstance(obj, dict):
        for k, v in obj.items():
            items += flatten(v, f"{prefix}.{k}" if prefix else k)
    elif isinstance(obj, list):
        for i, v in enumerate(obj):
            items += flatten(v, f"{prefix}.{i}")
    else:
        items.append((prefix, obj))
    return items


def get_path(obj, path):
    cur = obj
    for p in path.split("."):
        if cur is None:
            return ""
        cur = cur[int(p)] if isinstance(cur, list) else cur.get(p)
    return cur if cur is not None else ""


def style_header(ws, ncols):
    fill = PatternFill("solid", fgColor="1A2E5A")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
    ws.freeze_panes = "A2"


def main():
    mk = load(I18N / "mk.json")
    al = load(I18N / "al.json")
    en = load(I18N / "en.json")

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "UI text"
    ws.append(["key", "MK", "AL", "EN"])
    for key, _ in flatten(mk):
        ws.append([key, get_path(mk, key), get_path(al, key), get_path(en, key)])
    ws.column_dimensions["A"].width = 38
    for col in ("B", "C", "D"):
        ws.column_dimensions[col].width = 60
    style_header(ws, 4)

    parties = load(I18N / "parties.json")
    wp = wb.create_sheet("Parties")
    wp.append(["MK (party)", "AL", "EN", "Acronym MK", "Acronym AL", "Acronym EN"])
    for cyr, v in parties.items():
        wp.append([cyr, v["al"], v["en"], v["acrMk"], v["acrAl"], v["acrEn"]])
    wp.column_dimensions["A"].width = 52
    for col in ("B", "C"):
        wp.column_dimensions[col].width = 48
    style_header(wp, 6)

    names = load(I18N / "mp_names.json")
    wn = wb.create_sheet("MP names")
    wn.append(["MK (name)", "AL", "EN"])
    for cyr, v in names.items():
        wn.append([cyr, v["al"], v["en"]])
    for col in ("A", "B", "C"):
        wn.column_dimensions[col].width = 34
    style_header(wn, 3)

    # Two more simple MK | AL | EN maps: institutions and office case types.
    for sheet_name, fname, widths in [
        ("Institutions", "institutions.json", (52, 52, 52)),
        ("Case types", "case_types.json", (48, 48, 48)),
        ("Ethnicities", "ethnicities.json", (28, 28, 28)),
    ]:
        data = load(I18N / fname)
        wsx = wb.create_sheet(sheet_name)
        wsx.append(["MK", "AL", "EN"])
        for cyr, v in data.items():
            wsx.append([cyr, v["al"], v["en"]])
        for col, w in zip(("A", "B", "C"), widths):
            wsx.column_dimensions[col].width = w
        style_header(wsx, 3)

    # Questions + answers — machine-seeded by translate_questions.py, editable here.
    questions = json.loads((ROOT / "public" / "data" / "questions.json").read_text(encoding="utf-8"))
    qi18n_path = ROOT / "public" / "data" / "questions_i18n.json"
    qi18n = json.loads(qi18n_path.read_text(encoding="utf-8")) if qi18n_path.exists() else {}
    wq = wb.create_sheet("Questions")
    wq.append(["id", "MK question", "AL question", "EN question", "MK answer", "AL answer", "EN answer"])
    for q in questions:
        tr = qi18n.get(q["id"], {})
        qt, at = tr.get("q", {}), tr.get("a", {})
        mk_answer = "" if q.get("answerIsCopy") else (q.get("answer") or "")
        wq.append([q["id"], q.get("question") or "", qt.get("al", ""), qt.get("en", ""),
                   mk_answer, at.get("al", ""), at.get("en", "")])
    wq.column_dimensions["A"].width = 38
    for col in ("B", "C", "D", "E", "F", "G"):
        wq.column_dimensions[col].width = 60
    style_header(wq, 7)

    wb.save(OUT)
    print(f"✓ wrote {OUT.relative_to(ROOT)}  "
          f"(UI: {ws.max_row - 1}, Parties: {wp.max_row - 1}, MP names: {wn.max_row - 1}, "
          f"+ Institutions & Case types)")


if __name__ == "__main__":
    main()
